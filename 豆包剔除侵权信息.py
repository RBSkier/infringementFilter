import json
import re
import sys
from volcenginesdkarkruntime import Ark
from openpyxl import load_workbook

client = Ark(
    api_key= "1c89b646-1510-4379-b694-0ef845db8380", #os.environ.get("ARK_API_KEY"),
    # The base URL for model invocation 
    base_url="https://ark.cn-beijing.volces.com/api/v3",
    )

# 快速剔除内容（正则批量替换）
def fast_remove(exclude_list, original):
    if not exclude_list:
        return original
    escaped = [re.escape(item) for item in exclude_list]
    return re.compile("|".join(escaped)).sub("", original).strip()

# 配置参数（根据实际需求修改）
excel_path = "./产品翻译数据.xlsx"  # 你的Excel文件路径
sheet_name = "产品数据"  # 目标工作表名
start_row = 109  # 从第x行开始读取（Excel行号）
end_row = 1384

try:
    # 加载工作簿
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]  # 选择目标工作表
    # 逐行处理：从start_row开始读取，同时写入对应列
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        # 1. 读取当前行的B列（第2列）、C列（第3列）原始数据
        current_row_num = row[0].row
        original_title = []     # 存储B列数据
        original_description = []      # 存储C列数据
        original_title = row[1].value  # B列：原始标题
        original_desc = row[2].value    # C列：原始描述

        prompt = f"""完成任务：
                    1. 规则：从待处理文本中提取需要剔除的内容，剔除范围：
                        ①从标题文本中剔除小品牌名称，但知名品牌或者ip无需剔除；
                        ②从标题文本中剔除前后缀上的一些无含义乱码；
                        ③从描述文本中剔除小品牌名称、保修天数、质保期限相关描述；
                        ④从描述文本中剔除发货时效、full仓发货描述（美客多平台的包邮政策、售后相关无需剔除）；
                        ⑤从描述文本中剔除带有mercadolibre.com域名的链接的引流信息；
                    2. 处理对象：同时分析【标题文本】和【描述文本】，分别提取各自需要剔除的内容。
                    3. 原标题中剔除掉第1点的内容后，重新生成一个最多60个字母的标题，要求与剔除后的标题有差异化，避免被平台收录为跟卖链接。
                    3. 输出要求：仅返回标准JSON字典，无其他多余文字，字典格式如下：
                        {{
                            "title_exclude": [标题中需剔除的内容1, 标题中需剔除的内容2,...],
                            "description_exclude": [描述中需剔除的内容1, 描述中需剔除的内容2,...],
                            "new_title": "剔除内容新生成的标题"
                        }}
                    若某一文本无需要剔除的内容，对应列表为空数组[]。
                    4. 待处理文本：
                        【标题文本】：{original_title}
                        【描述文本】：{original_desc}"""

        #调用豆包处理数据
        try:
            completion = client.chat.completions.create(
                # Replace with Model ID
                model="doubao-seed-1-6-thinking-250715",
                messages=[
                    {"role": "user", "content": prompt}
                ],
                reasoning_effort='low'
            )
            response_text = completion.choices[0].message.content
        except Exception as api_e:
            print(f"❌ 第{current_row_num}行API调用失败：{api_e}，跳过当前行")
            continue  # 跳过当前行，继续处理下一行

        # 将豆包返回的剔除项转成列表类型
        title_exclude = []  # 标题剔除列表默认空
        description_exclude = []  # 描述剔除列表默认空
        new_title = ""  # 新标题默认空
        try:
            # 前置校验：过滤空和全空白字符串
            if response_text.strip():
                # 解析JSON字典
                exclude_dict = json.loads(response_text)
                # 提取对应列表（键不存在则赋空列表）
                title_exclude = exclude_dict.get("title_exclude", [])
                description_exclude = exclude_dict.get("description_exclude", [])
                new_title = exclude_dict.get("new_title", "")
        except json.JSONDecodeError as e:
            print(f"❌ JSON解析失败：{e}，使用空列表兜底")
            print(f"豆包实际返回内容：{response_text}")
            sys.exit(1)  # 终止程序，退出码1表示JSON解析异常
        except Exception as e:
            print(f"❌ 其他错误：{e}，使用空列表兜底")
            sys.exit(2)  # 终止程序，退出码2表示其他未知异常

        # 根据剔除项处理文本
        # clean_title = fast_remove(title_exclude, original_title)
        clean_description = fast_remove(description_exclude, original_desc)
        # 将记录回写excel的D、E、F、G列
        row[3].value = str(title_exclude)   # D列：title_exclude
        row[4].value = str(description_exclude) # E列：description_exclude
        row[5].value = new_title  # F列：new_title
        row[6].value = clean_description    # G列：clean_description
        row[7].value = completion.choices[0].message.reasoning_content    # H列：思考过程

        # 🔴 核心修改：每处理一行立即保存，防止数据丢失
        try:
            wb.save(excel_path)
            print(f"✅ 第{current_row_num}行处理完成，已保存到Excel")
        except PermissionError:
            print(f"❌ 第{current_row_num}行保存失败：Excel文件被手动打开，请关闭文件后重新运行")
            sys.exit(1)  # 文件被锁定时无法保存，终止程序
        except Exception as save_e:
            print(f"❌ 第{current_row_num}行保存失败：{save_e}，跳过当前行") 
    
    print(f"\n🎉 所有行处理完成！最终保存路径：{excel_path}")

except KeyError:
    print(f"❌ 工作表不存在：请检查工作表名「{sheet_name}」是否正确")
    sys.exit(1)
except Exception as main_e:
    print(f"❌ 程序主流程错误：{main_e}")
    sys.exit(1)
finally:
    # 关键：无论程序是否报错，都要手动关闭工作簿，避免文件损坏
    if wb is not None:
        wb.close()
        print(f"✅ Excel工作簿已正常关闭")